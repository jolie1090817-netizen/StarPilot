"""
敏感資料匿名化與對照表產生工具

功能：
1. 對 attend.xlsx 中的敏感資料（姓名、身分證字號）進行匿名化
2. 生成對照表用於還原
3. 保存匿名化後的 Excel 檔案
4. 輸出對照表到 CSV 檔案
"""

import openpyxl
import pandas as pd
import uuid
from pathlib import Path
from collections import defaultdict


class DataAnonymizer:
    def __init__(self, excel_file):
        """初始化匿名化工具"""
        self.excel_file = excel_file
        self.mapping = defaultdict(dict)  # 記錄映射關係
        self.load_workbook()

    def load_workbook(self):
        """載入 Excel 工作簿"""
        try:
            self.wb = openpyxl.load_workbook(self.excel_file)
            print(f"✅ 成功載入 {self.excel_file}")
            print(f"📊 包含 {len(self.wb.sheetnames)} 個工作表")
            print(f"📋 工作表列表: {', '.join(self.wb.sheetnames)}")
        except Exception as e:
            print(f"❌ 錯誤：無法載入檔案 - {e}")
            raise

    def generate_fake_name(self, original_name):
        """根據原始姓名生成假名"""
        if original_name not in self.mapping['names']:
            # 生成格式：假名_XXXX（4位隨機數字）
            random_suffix = str(uuid.uuid4().hex[:4]).upper()
            fake_name = f"員工_{random_suffix}"
            self.mapping['names'][original_name] = fake_name
        return self.mapping['names'][original_name]

    def generate_fake_id(self, original_id):
        """根據原始身分證號生成假號碼"""
        if original_id not in self.mapping['ids']:
            # 生成格式：偽_XXXXXXXXXXXX（12個隨機字符）
            random_suffix = uuid.uuid4().hex[:12].upper()
            fake_id = f"ID_{random_suffix}"
            self.mapping['ids'][original_id] = fake_id
        return self.mapping['ids'][original_id]

    def find_sensitive_columns(self, ws):
        """找出包含敏感資料的欄位"""
        sensitive_cols = {}

        # 掃描第一行找出相關欄位
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                col_name = str(cell.value).lower()
                # 尋找姓名相關欄位
                if any(keyword in col_name for keyword in ['姓名', 'name', '名字']):
                    sensitive_cols['name'] = col_idx
                # 尋找身分證相關欄位
                elif any(keyword in col_name for keyword in ['身分證', 'id', '身份證', '証号']):
                    sensitive_cols['id'] = col_idx

        return sensitive_cols

    def anonymize_sheet(self, sheet_name):
        """匿名化單個工作表"""
        ws = self.wb[sheet_name]

        # Find sensitive columns
        sensitive_cols = self.find_sensitive_columns(ws)

        if not sensitive_cols:
            print(f"  ⚠️  {sheet_name}: 未找到敏感資料欄位")
            return 0

        anonymized_count = 0

        # Anonymize data rows (skip header)
        for row in ws.iter_rows(min_row=2):
            # 處理姓名欄位
            if 'name' in sensitive_cols:
                col_idx = sensitive_cols['name']
                original_name = row[col_idx - 1].value
                if original_name:
                    fake_name = self.generate_fake_name(original_name)
                    row[col_idx - 1].value = fake_name
                    anonymized_count += 1

            # 處理身分證欄位
            if 'id' in sensitive_cols:
                col_idx = sensitive_cols['id']
                original_id = row[col_idx - 1].value
                if original_id:
                    fake_id = self.generate_fake_id(original_id)
                    row[col_idx - 1].value = fake_id
                    anonymized_count += 1

        return anonymized_count

    def process_all_sheets(self):
        """處理所有工作表"""
        print("\n🔄 開始匿名化流程...")
        total_anonymized = 0

        for sheet_name in self.wb.sheetnames:
            count = self.anonymize_sheet(sheet_name)
            if count > 0:
                print(f"  ✅ {sheet_name}: 已匿名化 {count} 個敏感資料")
            total_anonymized += count

        print(f"\n✅ 總共匿名化 {total_anonymized} 個敏感資料項目")
        return total_anonymized

    def save_anonymized_file(self, output_path):
        """保存匿名化後的 Excel 檔案"""
        try:
            self.wb.save(output_path)
            print(f"💾 已保存匿名化檔案: {output_path}")
        except Exception as e:
            print(f"❌ 儲存失敗: {e}")
            raise

    def create_mapping_table(self, output_path):
        """建立對照表"""
        mapping_data = []

        # 姓名對照表
        for original_name, fake_name in sorted(self.mapping['names'].items()):
            mapping_data.append({
                '原始資料': original_name,
                '匿名化後': fake_name,
                '資料類型': '姓名',
                '還原方式': f"將 '{fake_name}' 替換為 '{original_name}'"
            })

        # 身分證對照表
        for original_id, fake_id in sorted(self.mapping['ids'].items()):
            mapping_data.append({
                '原始資料': original_id,
                '匿名化後': fake_id,
                '資料類型': '身分證字號',
                '還原方式': f"將 '{fake_id}' 替換為 '{original_id}'"
            })

        # 保存為 CSV
        df = pd.DataFrame(mapping_data)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"📋 已生成對照表: {output_path}")
        print(f"   共 {len(mapping_data)} 筆對照記錄")

        return df

    def display_mapping_summary(self):
        """顯示對照表摘要"""
        print("\n" + "="*60)
        print("📊 對照表摘要")
        print("="*60)

        if self.mapping['names']:
            print(f"\n👤 姓名對照 ({len(self.mapping['names'])} 筆):")
            for original, fake in list(self.mapping['names'].items())[:5]:
                print(f"   {original} → {fake}")
            if len(self.mapping['names']) > 5:
                print(f"   ... 及其他 {len(self.mapping['names']) - 5} 筆")

        if self.mapping['ids']:
            print(f"\n🆔 身分證對照 ({len(self.mapping['ids'])} 筆):")
            for original, fake in list(self.mapping['ids'].items())[:5]:
                print(f"   {original} → {fake}")
            if len(self.mapping['ids']) > 5:
                print(f"   ... 及其他 {len(self.mapping['ids']) - 5} 筆")

        print("\n" + "="*60)


def main():
    """主程式"""
    # 設定路徑
    base_dir = Path(__file__).parent
    excel_file = base_dir / '測試資料_待整理' / 'attend.xlsx'
    output_dir = base_dir / '匿名化結果'

    # 建立輸出目錄
    output_dir.mkdir(exist_ok=True)

    anonymized_file = output_dir / 'attend_anonymized.xlsx'
    mapping_file = output_dir / '對照表_身分還原用.csv'

    print("🚀 開始執行敏感資料匿名化程序")
    print(f"📁 來源檔案: {excel_file}")
    print(f"📁 輸出目錄: {output_dir}")
    print()

    # 檢查檔案是否存在
    if not excel_file.exists():
        print(f"❌ 錯誤：找不到 {excel_file}")
        return

    # 執行匿名化
    anonymizer = DataAnonymizer(excel_file)
    anonymizer.process_all_sheets()

    # 顯示對照表摘要
    anonymizer.display_mapping_summary()

    # 保存檔案
    anonymizer.save_anonymized_file(anonymized_file)
    anonymizer.create_mapping_table(mapping_file)

    print("\n✨ 程序完成！")
    print(f"\n📌 重要提醒：")
    print(f"   1. 對照表文件已保存至: {mapping_file}")
    print(f"   2. 請妥善保管對照表，用於日後資料還原")
    print(f"   3. 建議保存對照表的備份副本")


if __name__ == '__main__':
    main()
