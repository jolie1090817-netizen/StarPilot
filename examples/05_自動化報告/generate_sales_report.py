#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
📊 自動化銷售報告生成系統
功能:
- 讀取 CSV 銷售數據
- 自動生成統計分析
- 建立美化的 Excel 報告含圖表
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
import os

# ============ 設定 ============
DATA_FILE = "../../data/sales.csv"
OUTPUT_DIR = "reports"
OUTPUT_FILE = f"週報_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

# ============ 確保輸出目錄存在 ============
Path(OUTPUT_DIR).mkdir(exist_ok=True)

# ============ 讀取數據 ============
print("📖 讀取銷售數據...")
df = pd.read_csv(DATA_FILE)
df['日期'] = pd.to_datetime(df['日期'])

print(f"✅ 數據已載入，共 {len(df)} 筆記錄")
print(f"   時間範圍: {df['日期'].min().date()} ~ {df['日期'].max().date()}")

# ============ 統計分析 ============
print("\n📊 生成統計分析...")

# 每個產品的統計
product_stats = df.groupby('產品').agg({
    '數量': 'sum',
    '金額': 'sum'
}).round(0).astype(int)
product_stats = product_stats.sort_values('金額', ascending=False)

# 按日期的銷售趨勢
daily_sales = df.groupby('日期')['金額'].sum().reset_index()

# 算出前 5 名產品
top_5_products = product_stats.head(5)

print(f"   - 共 {len(product_stats)} 個產品")
print(f"   - 總銷售額: NT${df['金額'].sum():,}")
print(f"   - 平均每筆: NT${df['金額'].mean():.0f}")

# ============ 建立 Excel 報告 ============
print("\n📝 建立 Excel 報告...")

with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    # 🔹 工作表 1: 原始數據
    df_export = df.copy()
    df_export['日期'] = df_export['日期'].astype(str)
    df_export.to_excel(writer, sheet_name='原始數據', index=False)

    # 🔹 工作表 2: 統計摘要
    summary_data = {
        '統計項目': ['總銷售額', '總銷售量', '平均單筆', '產品數', '交易日期'],
        '數值': [
            f"NT${df['金額'].sum():,}",
            f"{df['數量'].sum()} 件",
            f"NT${df['金額'].mean():.0f}",
            f"{len(product_stats)} 個",
            f"{len(df['日期'].dt.date.unique())} 天"
        ]
    }
    pd.DataFrame(summary_data).to_excel(writer, sheet_name='統計摘要', index=False)

    # 🔹 工作表 3: 產品排行
    product_stats.to_excel(writer, sheet_name='產品排行')

    # 🔹 工作表 4: 銷售趨勢
    daily_sales.to_excel(writer, sheet_name='銷售趨勢', index=False)

# ============ 加入圖表 ============
print("📈 建立圖表...")

wb = openpyxl.load_workbook(OUTPUT_FILE)

# 圖表 1: 前 5 名產品銷售額 (長條圖)
ws_chart1 = wb.create_sheet('圖表_產品排行', 0)
top_5_data = product_stats.head(5)
# 手動寫入標題
ws_chart1['A1'] = '產品'
ws_chart1['B1'] = '數量'
ws_chart1['C1'] = '金額'
for idx, (product, row) in enumerate(top_5_data.iterrows(), start=2):
    ws_chart1[f'A{idx}'] = product
    ws_chart1[f'B{idx}'] = row['數量']
    ws_chart1[f'C{idx}'] = row['金額']

# 長條圖
bar_chart = BarChart()
bar_chart.type = "col"
bar_chart.title = "前 5 名產品銷售額"
bar_chart.y_axis.title = "銷售額 (NT$)"
bar_chart.x_axis.title = "產品"
labels = Reference(ws_chart1, min_col=1, min_row=2,
                   max_row=len(top_5_data) + 1)
data = Reference(ws_chart1, min_col=3, min_row=1, max_row=len(top_5_data) + 1)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(labels)
bar_chart.height = 12
bar_chart.width = 20
ws_chart1.add_chart(bar_chart, "E2")

# 圖表 2: 產品佔比 (圓餅圖)
ws_chart2 = wb.create_sheet('圖表_銷售佔比')
ws_chart2['A1'] = '產品'
ws_chart2['B1'] = '數量'
ws_chart2['C1'] = '金額'
for idx, (product, row) in enumerate(product_stats.iterrows(), start=2):
    ws_chart2[f'A{idx}'] = product
    ws_chart2[f'B{idx}'] = row['數量']
    ws_chart2[f'C{idx}'] = row['金額']

pie_chart = PieChart()
pie_chart.title = "各產品銷售佔比"
labels = Reference(ws_chart2, min_col=1, min_row=2,
                   max_row=len(product_stats) + 1)
data = Reference(ws_chart2, min_col=3, min_row=1,
                 max_row=len(product_stats) + 1)
pie_chart.add_data(data, titles_from_data=True)
pie_chart.set_categories(labels)
pie_chart.height = 12
pie_chart.width = 18
ws_chart2.add_chart(pie_chart, "E2")

# 圖表 3: 銷售趨勢 (折線圖)
ws_chart3 = wb.create_sheet('圖表_銷售趨勢')
ws_chart3['A1'] = '日期'
ws_chart3['B1'] = '金額'
for idx, row in enumerate(daily_sales.iterrows(), start=2):
    ws_chart3[f'A{idx}'] = str(row[1]['日期'].date())
    ws_chart3[f'B{idx}'] = row[1]['金額']

line_chart = LineChart()
line_chart.title = "每日銷售趨勢"
line_chart.y_axis.title = "銷售額 (NT$)"
line_chart.x_axis.title = "日期"
labels = Reference(ws_chart3, min_col=1, min_row=2,
                   max_row=len(daily_sales) + 1)
data = Reference(ws_chart3, min_col=2, min_row=1, max_row=len(daily_sales) + 1)
line_chart.add_data(data, titles_from_data=True)
line_chart.set_categories(labels)
line_chart.height = 12
line_chart.width = 20
ws_chart3.add_chart(line_chart, "E2")

# ============ 美化 Excel 格式 ============
print("🎨 格式化 Excel...")

# 定義樣式
header_fill = PatternFill(start_color="4472C4",
                          end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
center_align = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 格式化統計摘要
ws_summary = wb['統計摘要']
for col in ['A', 'B']:
    for row in ws_summary.iter_rows(min_col=ord(col)-64, max_col=ord(col)-64):
        for cell in row:
            if cell.row == 1:  # 標題行
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            cell.border = border

# 自動調整列寬
for ws_name in wb.sheetnames:
    sheet = wb[ws_name]
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

# 保存
wb.save(OUTPUT_FILE)

# ============ 輸出完成信息 ============
print(f"\n✨ 報告生成完成！")
print(f"📁 保存位置: {os.path.abspath(OUTPUT_FILE)}")
print(f"📊 包含以下內容:")
print(f"   ✓ 原始數據工作表")
print(f"   ✓ 統計摘要工作表")
print(f"   ✓ 產品排行工作表")
print(f"   ✓ 銷售趨勢工作表")
print(f"   ✓ 前 5 名產品長條圖")
print(f"   ✓ 產品銷售佔比圓餅圖")
print(f"   ✓ 每日銷售趨勢折線圖")
print(f"\n🎉 下次只需執行此腳本即可自動生成最新報告！")
